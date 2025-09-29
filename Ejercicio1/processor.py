import os
import pandas as pd
from openpyxl import Workbook
from pathlib import Path

base_dir = r"D:\USJ_2025_2026\Ingenieria de Datos\Git\ingenieria_de_datos\Ejercicio1"
download_dir = os.path.join(base_dir, "downloads")
excel_dir = os.path.join(base_dir, "processor")
processed_dir = os.path.join(base_dir, "processed")
os.makedirs(excel_dir, exist_ok=True)
os.makedirs(processed_dir, exist_ok=True)

#cargar CSVs como (nombre, df) ---
frames = []
for csv_name in os.listdir(download_dir):
    if csv_name.lower().endswith(".csv"):
        path = os.path.join(download_dir, csv_name)
        df = pd.read_csv(path)
        name = Path(csv_name).stem
        frames.append((name, df))

#función simple: media de duración en MINUTOS según 3 casos ---
def mean_trip_minutes(df: pd.DataFrame) -> float | None:
    #'tripduration' (s)
    if "tripduration" in df.columns:
        s = pd.to_numeric(df["tripduration"], errors="coerce") / 60.0
        return float(s.mean(skipna=True))

    #'01 - Rental Details Duration In Seconds Uncapped' (s)
    col_alt = "01 - Rental Details Duration In Seconds Uncapped"
    if col_alt in df.columns:
        s = pd.to_numeric(df[col_alt], errors="coerce") / 60.0
        return float(s.mean(skipna=True))

    #ended_at - started_at (min)
    if {"started_at", "ended_at"}.issubset(df.columns):
        start = pd.to_datetime(df["started_at"], errors="coerce")
        end   = pd.to_datetime(df["ended_at"], errors="coerce")
        mins = (end - start).dt.total_seconds() / 60.0
        mins = mins.where(mins >= 0)  # evita negativas si hay registros invertidos
        return float(pd.Series(mins).mean(skipna=True))

    #si no encaja ningún caso
    return None

#escribir nombres en Excel (col A) y medias en col B
wb = Workbook()
ws = wb.active
ws.title = "Resumen"
ws["A1"] = "dataset"
ws["B1"] = "media_duracion_min"

results = []  #para CSV en processed/
row = 2
for name, df in frames:
    mean_min = mean_trip_minutes(df)
    ws.cell(row=row, column=1, value=name)
    ws.cell(row=row, column=2, value=mean_min)
    results.append({"dataset": name, "mean_duracion_min": mean_min})
    row += 1

#guardar Excel
out_path_xlsx = os.path.join(excel_dir, "excel.xlsx")
wb.save(out_path_xlsx)
print("Excel guardado en:", out_path_xlsx)

#guardar resumen también en processed/ como CSV
out_path_csv = os.path.join(processed_dir, "mean_trip_time.csv")
pd.DataFrame(results).sort_values("dataset").to_csv(out_path_csv, index=False)
print("CSV guardado en:", out_path_csv)
